"""mixed_pack_smoke.py — one-command release gate for the mixed-pack sprint (lane 04 / J).

Run from repo root:  python -m engine.scripts.mixed_pack_smoke

Phase A (runs today, no backend parser needed): every format in a pack — PDF, DOCX, XLSX, CSV — parses,
and the deterministic net (engine.gating_scan) flags the planted deal-breakers in each. This is the
"no silent deal-breaker miss, even in a spreadsheet" invariant, checked on the real fixtures/mixed-pack/ files.

Phase B (auto-activates once backend lane 01 lands): probes backend.app.ingest for a non-PDF ingest
entrypoint; if present, reads each Office fixture through the BACKEND and re-checks the net end-to-end.
Skips with a clear PENDING note (still exit 0) until the parser exists.

Phase C (J-096 — ZIP support): unzips fixtures/mixed-pack/sample-pack.zip (a DOCX+XLSX+CSV pack, plus
a __MACOSX junk entry and an unsupported notes.txt to prove clean skipping) through the backend's own
_expand_zip(), same code path POST /tenders/upload uses, then runs the net over each extracted entry.
Skips with a clear PENDING note until the backend exposes ZIP support.

Exit 0 = safe so far. Non-zero = a real regression (a fixture stopped parsing, or the net stopped catching
a planted gate, or the backend read a file but broke the invariant).
"""
from __future__ import annotations

import csv as _csv
from pathlib import Path

from engine.gating_scan import scan_candidates

REPO = Path(__file__).resolve().parents[2]
FIX = REPO / "fixtures" / "mixed-pack"
PDF_SAMPLE = REPO / "frontend" / "public" / "demo" / "bradwell-grounds-itt.pdf"
ZIP_SAMPLE = FIX / "sample-pack.zip"


def _gates(text: str) -> int:
    return len(scan_candidates([(1, text)]))


# ---- format readers (Phase A reads the fixtures directly; Phase B uses the backend) ----
def _docx_text(path: Path) -> str:
    from docx import Document
    return "\n".join(p.text for p in Document(str(path)).paragraphs)


def _xlsx_text(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True)
    try:
        return "\n".join(
            str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value
        )
    finally:
        wb.close()  # read_only keeps the file mapped open otherwise (breaks cleanup on Windows)


def _csv_text(path: Path) -> str:
    with open(path, encoding="utf-8") as f:
        return "\n".join(" ".join(row) for row in _csv.reader(f))


def _pdf_text(path: Path) -> str:
    from backend.app.ingest import ingest_pdf
    doc = ingest_pdf(path, enrich=False)
    return "\n".join(p.text for p in doc.pages)


FIXTURES = [
    ("PDF   (sample ITT)", PDF_SAMPLE, _pdf_text),
    ("DOCX  (return forms)", FIX / "sample-return-forms.docx", _docx_text),
    ("XLSX  (pricing sched)", FIX / "sample-pricing-schedule.xlsx", _xlsx_text),
    ("CSV   (compliance)", FIX / "sample-compliance.csv", _csv_text),
]


def phase_a() -> bool:
    print("== Phase A — format-neutral deal-breaker net (no backend needed) ==")
    ok = True
    for label, path, reader in FIXTURES:
        if not path.exists():
            print(f"  FAIL  {label}: missing file {path}")
            ok = False
            continue
        try:
            gates = _gates(reader(path))
        except ImportError as e:
            print(f"  SKIP  {label}: {e} (pip install python-docx openpyxl)")
            continue
        except Exception as e:  # noqa: BLE001 - a smoke tool reports, doesn't raise
            print(f"  FAIL  {label}: could not read ({type(e).__name__}: {e})")
            ok = False
            continue
        mark = "OK  " if gates > 0 else "FAIL"
        if gates == 0:
            ok = False
        print(f"  {mark}  {label}: net flagged {gates} disqualifier line(s)")
    return ok


def _backend_pack_ingest():
    """Return a callable(path)->IngestedDoc for non-PDF files once lane 01 exposes one, else None.
    Contract lane 01 must satisfy: a function on backend.app.ingest that turns a .docx/.xlsx/.csv path
    into an IngestedDoc (`.pages`, each with `.number` + `.text`)."""
    try:
        from backend.app import ingest as bi
    except Exception:  # noqa: BLE001
        return None
    for name in ("ingest_file", "ingest_pack", "ingest_any", "ingest_document"):
        fn = getattr(bi, name, None)
        if callable(fn):
            return fn
    return None


def phase_b() -> bool:
    print("\n== Phase B — backend reads the pack (auto-activates when lane 01 lands) ==")
    fn = _backend_pack_ingest()
    if fn is None:
        print("  PENDING  backend has no non-PDF ingest entrypoint yet (lane 01 —")
        print("           expose ingest_file/ingest_pack(path)->IngestedDoc; this gate then runs).")
        return True  # expected before the parser lands — not a failure
    ok = True
    for label, path, _ in FIXTURES:
        if path.suffix.lower() == ".pdf" or not path.exists():
            continue
        try:
            doc = fn(str(path))
            text = "\n".join(p.text for p in doc.pages)
            gates = _gates(text)
            mark = "OK  " if (text.strip() and gates > 0) else "FAIL"
            if not (text.strip() and gates > 0):
                ok = False
            print(f"  {mark}  {label}: backend read it, net flagged {gates} gate(s)")
        except Exception as e:  # noqa: BLE001
            print(f"  FAIL  {label}: backend ingest raised ({type(e).__name__}: {e})")
            ok = False
    return ok


_READERS_BY_EXT = {".docx": _docx_text, ".xlsx": _xlsx_text, ".csv": _csv_text}


def phase_c() -> bool:
    print("\n== Phase C — ZIP pack support (J-096) ==")
    if not ZIP_SAMPLE.exists():
        print(f"  FAIL  missing fixture {ZIP_SAMPLE}")
        return False
    try:
        from backend.app.main import _expand_zip
    except ImportError:
        print("  PENDING  backend has no ZIP expansion yet (J-096 — expose _expand_zip in")
        print("           backend.app.main; this gate then runs).")
        return True  # expected before ZIP support lands — not a failure
    raw = ZIP_SAMPLE.read_bytes()
    entries = _expand_zip(raw, ZIP_SAMPLE.name)
    ok = True
    names = {name for name, _ in entries}
    if "notes.txt" in names:
        print("  FAIL  unsupported entry 'notes.txt' was not skipped")
        ok = False
    for name, content in entries:
        ext = Path(name).suffix.lower()
        reader = _READERS_BY_EXT.get(ext)
        if reader is None:
            print(f"  FAIL  unexpected entry extracted: {name}")
            ok = False
            continue
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)
        try:
            gates = _gates(reader(tmp_path))
        except Exception as e:  # noqa: BLE001
            print(f"  FAIL  {name}: could not read extracted entry ({type(e).__name__}: {e})")
            ok = False
            continue
        finally:
            tmp_path.unlink(missing_ok=True)
        mark = "OK  " if gates > 0 else "FAIL"
        if gates == 0:
            ok = False
        print(f"  {mark}  {name}: extracted from ZIP, net flagged {gates} disqualifier line(s)")
    if names != {"sample-return-forms.docx", "sample-pricing-schedule.xlsx", "sample-compliance.csv"}:
        print(f"  FAIL  unexpected entry set extracted: {sorted(names)}")
        ok = False
    return ok


def main() -> int:
    a = phase_a()
    b = phase_b()
    c = phase_c()
    print("\n== Still eyeballed by the release owner (not automated here) ==")
    for item in (
        "every requirement carries source_filename (its origin file)",
        "non-PDF rows show an excerpt + locator (page/sheet/row)",
        "NO fake PDF highlight appears for Office-derived rows",
        "unsupported files fail clearly (not silently)",
    ):
        print(f"  [ ] {item}")
    ok = a and b and c
    print("\n" + ("SMOKE OK — safe so far." if ok else "SMOKE FAILED — a real regression, do not ship."))
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
